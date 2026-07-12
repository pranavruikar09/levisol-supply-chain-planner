"""Levisol Planning System - Excel report generation (openpyxl via pandas)."""
import io
import numpy as np
import pandas as pd


def _fmt(ws_writer, df, sheet, freeze="A4", title=None):
    df.to_excel(ws_writer, sheet_name=sheet, index=False, startrow=2)
    wb = ws_writer.book
    ws = ws_writer.sheets[sheet]
    from openpyxl.styles import Font, PatternFill, Alignment
    ws["A1"] = title or sheet
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[3]:
        if cell.value is not None:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = hdr_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col[2:20] if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 10), 28)
    ws.freeze_panes = freeze


def norms_workbook(norms, hub_norms_df, tiers, xyz, fe, path=None):
    buf = path or io.BytesIO()
    t = tiers.reset_index().rename(columns={"index": "sku"})
    t["xyz"] = t["sku"].map(xyz)
    t["abc_xyz"] = t["tier"] + "-" + t["xyz"]
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        _fmt(w, norms, "CFA Norms", title="Inventory Norms by SKU x CFA (safety stock, ROP, days of cover)")
        _fmt(w, hub_norms_df, "Hub Norms", title="Inventory Norms by SKU x Hub (98% service, risk-pooled)")
        _fmt(w, t, "SKU Tiers", title="SKU Tier Classification (ABC by 50/30/15/5 volume slabs) + XYZ volatility")
        _fmt(w, fe.round(3), "Forecast Accuracy", title="Forecast Error Metrics by SKU x CFA (bias, MAD, RMSE, MAPE)")
    return buf


def plan_workbook(plan, data, tiers, path=None):
    buf = path or io.BytesIO()
    cost_df = pd.DataFrame([
        {"Cost head": "Production", "Rs": plan.costs["production"]},
        {"Cost head": "Freight: plant -> hub", "Rs": plan.costs["transport_plant_hub"]},
        {"Cost head": "Freight: hub -> CFA", "Rs": plan.costs["transport_hub_cfa"]},
        {"Cost head": "Penalty: unmet demand", "Rs": plan.costs["penalty_unmet"]},
        {"Cost head": "Penalty: hub SS shortfall", "Rs": plan.costs["hub_ss_shortfall"]},
        {"Cost head": "TOTAL", "Rs": plan.costs["total"]},
    ])
    cost_df["Rs crore"] = (cost_df["Rs"] / 1e7).round(3)
    prod = plan.production.copy()
    prod["tier"] = prod["sku"].map(tiers["tier"])
    prod = prod.sort_values(["plant", "line", "sku"])
    util = pd.DataFrame([{"plant_line": k, "utilization_pct": round(v * 100, 1)}
                         for k, v in plan.kpis["utilization"].items()])
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        _fmt(w, cost_df, "Cost Summary", title="Jan-2026 Plan - Total Cost Breakdown")
        _fmt(w, prod, "Production Plan", title="Production by SKU x Plant (25 kL batches)")
        _fmt(w, plan.plant_hub.round(2), "Plant-Hub Flows", title="Volume routed Plant -> Hub (kL)")
        _fmt(w, plan.hub_cfa.round(3), "Hub-CFA Dispatch", title="Dispatch Hub -> CFA by SKU (kL)")
        _fmt(w, plan.hub_position, "Hub Stock Position", title="Hub opening vs safety-stock target vs ending stock")
        _fmt(w, plan.unmet if len(plan.unmet) else pd.DataFrame([{"note": "No unmet demand"}]),
             "Unmet Demand", title="Demand not served - what, how much, cost")
        _fmt(w, util, "Utilization", title="Plant x Line Utilization")
    return buf


# ===================================================================== V2 additions
def exec_summary_workbook(plan, data, tiers, recs, constraints, path=None):
    """One-page executive summary + recommendations + constraint compliance."""
    import pandas as pd, io as _io
    buf = path or _io.BytesIO()
    k = plan.kpis
    summary = pd.DataFrame([
        ["Total plan cost", f"₹{plan.costs['total']/1e7:,.2f} crore"],
        ["  Production", f"₹{plan.costs['production']/1e7:,.2f} cr"],
        ["  Freight (both legs)", f"₹{(plan.costs['transport_plant_hub']+plan.costs['transport_hub_cfa'])/1e7:,.2f} cr"],
        ["  Penalties (unmet + hub SS)", f"₹{(plan.costs['penalty_unmet']+plan.costs['hub_ss_shortfall'])/1e7:,.2f} cr"],
        ["Fill rate", f"{k['fill_rate']*100:,.2f}%"],
        ["Demand served", f"{k['served_kl']:,.0f} of {k['total_demand_kl']:,.0f} kL"],
        ["Production volume", f"{k['total_production_kl']:,.0f} kL "
                              f"({int(plan.production['batches'].sum()) if len(plan.production) else 0} batches)"],
        ["Hub SS shortfall", f"{plan.hub_position['shortfall_kl'].sum():,.1f} kL"],
    ], columns=["KPI", "Value"])
    recs_df = pd.DataFrame([{"Signal": r["icon"], "Recommendation": r["title"],
                             "Detail": r["detail"]} for r in recs])
    cons_df = pd.DataFrame(constraints)
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        _fmt(w, summary, "Executive Summary", title="Levisol Monthly Plan — Executive Summary")
        _fmt(w, recs_df, "Recommendations", title="Planner Recommendations")
        _fmt(w, cons_df, "Constraints", title="Optimization Constraint Compliance")
    return buf


def scenario_comparison_workbook(scenarios: dict, path=None):
    """scenarios: {name: result-dict from core.scenario.run_scenario}"""
    import pandas as pd, io as _io
    buf = path or _io.BytesIO()
    rows, cost_rows = [], []
    for name, r in scenarios.items():
        s, c = r["summary"], r["plan"].costs
        rows.append({"Scenario": name,
                     "Total cost (₹ cr)": round(s["total_cost"] / 1e7, 3),
                     "Fill rate %": round(s["fill_rate"] * 100, 2),
                     "Production kL": round(s["production_kl"], 0),
                     "Unmet kL": round(s["unmet_kl"], 1),
                     "Inventory norm kL": round(s["inventory_norm_kl"], 0)})
        cost_rows.append({"Scenario": name,
                          **{k2.replace("_", " "): round(v / 1e7, 3) for k2, v in c.items()}})
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        _fmt(w, pd.DataFrame(rows), "KPI Comparison", title="Scenario KPI Comparison")
        _fmt(w, pd.DataFrame(cost_rows), "Cost Comparison", title="Scenario Cost Heads (₹ crore)")
    return buf
